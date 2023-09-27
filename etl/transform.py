from difflib import get_close_matches
from glob import glob
from pathlib import Path
from pathlib import PosixPath

import pandas as pd
import toml
import xlrd
from loguru import logger
from openpyxl import load_workbook
from parse import parse

GEOGRAPHY_CODE_LEN = 9  # Length of geo code, eg E06000047

# Read local `config.toml` file.
config = toml.load("config.toml")
MIN_YEAR = config["min_year"]


def get_sheetnames_xlsx(file_path: PosixPath) -> list[str]:
    """Given an xlsx file, returns the names of the sheets in it.

    Args:
        file_path (PosixPath): path to the xlsx file

    Returns:
        list[str]: A list of the worksheet names
    """
    wb = load_workbook(file_path, read_only=True, keep_links=False)
    return wb.sheetnames


def get_sheetnames_xls(filepath: PosixPath) -> list[str]:
    """Given an xls file, returns the names of the sheets in it.

    Args:
        filepath (PosixPath): path to the xls file

    Returns:
        list[str]: A list of the worksheet names
    """
    xls = xlrd.open_workbook(filepath, on_demand=True)
    return xls.sheet_names()


def get_files_list() -> pd.DataFrame:
    # parse files
    file_list = glob(config["downloads_location"] + "[0-9][0-9][0-9][0-9].xls*")
    text_to_parse = config["downloads_location"] + "{year}.{file_extension}"
    list_of_dict_of_info = [
        parse(text_to_parse, file_path).named for file_path in file_list
    ]
    for x, y in zip(list_of_dict_of_info, file_list):
        x["file_path"] = Path(y)
    df_meta = pd.DataFrame.from_dict(list_of_dict_of_info)
    df_meta["year"] = df_meta["year"].astype(int)
    df_meta = df_meta.loc[df_meta["year"] >= MIN_YEAR].copy()
    df_meta = df_meta.reset_index(drop=True)
    return df_meta


def add_sheet_list_to_df_meta(df_meta: pd.DataFrame) -> pd.DataFrame:
    df_meta["sheet_names"] = "None"
    for file_ext, sheet_func in zip(
        ["xlsx", "xls"], [get_sheetnames_xlsx, get_sheetnames_xls]
    ):
        df_meta.loc[df_meta["file_extension"] == file_ext, "sheet_names"] = df_meta.loc[
            df_meta["file_extension"] == file_ext, :
        ].apply(lambda x: sheet_func(x["file_path"]), axis=1)
    return df_meta


def nominate_relevant_worksheet(worksheet_names: list[str]) -> str:
    # special case where "1" is in list
    if "1" in worksheet_names:
        return "1"
    else:
        return get_close_matches("Figures", worksheet_names, n=1)[0]


def open_file_and_clean_it(file_path: PosixPath, worksheet: str, year: int):
    logger.info(f"Running open_file_and_clean_it on {file_path}")
    df = pd.read_excel(file_path, sheet_name=worksheet)
    # remove rows with more than a few nans across
    df = df.dropna(axis=0, thresh=3)
    df = df.reset_index(drop=True)
    # make first row the header
    df.columns = df.iloc[0]
    df = df.iloc[1:].copy().reset_index(drop=True)
    # drop any row with mixed geographies (assuming geography is 1st column)
    df = df[df.iloc[:, 0].str.len() == GEOGRAPHY_CODE_LEN].copy()
    # nice column names
    # detect how many months of data there are
    num_months = df.shape[1] - 2
    date_range = pd.date_range(start=f"{year-1}/1/1", freq="m", periods=num_months)
    df.columns = ["geo_code", "place_name"] + list(
        date_range.strftime("%B").str.lower()
    )
    # tidy up place names
    df["place_name"] = df["place_name"].str.lower()
    # make data tidy
    df = pd.melt(
        df, id_vars=["geo_code", "place_name"], var_name="month", value_name="deaths"
    )
    df = df.astype(
        {
            "geo_code": "string[pyarrow]",
            "place_name": "string[pyarrow]",
            "month": "string[pyarrow]",
            "deaths": "float32[pyarrow]",
        }
    )
    return df


def transform_from_excel_to_tidy_parquet() -> None:
    df_meta = get_files_list()
    df_meta = add_sheet_list_to_df_meta(df_meta)
    df_meta["worksheet"] = df_meta["sheet_names"].apply(nominate_relevant_worksheet)

    df_tidy = pd.DataFrame()
    for index, row in df_meta.iterrows():
        df = open_file_and_clean_it(row["file_path"], row["worksheet"], row["year"])
        df["year"] = row["year"]
        df_tidy = pd.concat([df, df_tidy], axis=0)

    df_tidy["datetime"] = (
        pd.to_datetime(
            df_tidy["year"].astype("string") + "-" + df_tidy["month"], format="%Y-%B"
        )
        + pd.offsets.BMonthEnd()
    )
    df_tidy.to_parquet(Path(config["downloads_location"]) / config["name_of_data_file"])


if __name__ == "__main__":
    transform_from_excel_to_tidy_parquet()
